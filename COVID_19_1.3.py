from bs4 import BeautifulSoup as bf
import requests as rq
import pymysql
import time
import pickle
import openpyxl
import os
import re
import json

def access_web(name,url):
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}
    # url = 'https://www.worldometers.info/coronavirus/'
    file_name = name + '_疫情数据.txt'
    print('accessing web: %s'%url)
    response = rq.get(url=url, headers=headers)
    # print('analyzing data')
    # today = '_'.join([str('%02d'%_) for _ in time.localtime()[:3]])
    with open(file_name, 'w', encoding='utf-8') as f:
        f.write(response.text)
    return response

def country_data(tag):
    """
    return a list "info" contains the info listed as follows:

    info[0] = "country, other"
    info[1] = "tatal cases"
    info[2] = "new cases"
    info[3] = "total death"
    info[4] = "new death"
    info[5] = "total recovered"
    info[6] = "avtive cases"
    info[7] = "serious, critical"

    info[8] = "tot cases/1M pop"
    info[9] = "death/1M pop"
    info[10] = "1st case"
    """
    info = tag.text.split('\n')
    info.pop(0)
    info.pop()
    return info

def get_target(web):
    soup = bf(web, 'lxml')
    tbody = soup.find('tbody')
    trs = tbody.find_all(name='tr', class_=False)
    return trs

def formatdata(statistic):
    for i in range(len(statistic)):
        if statistic[i] == '' or statistic[i] == ' ':
            statistic[i] = '0'
        elif statistic[i] == 'N/A':
            statistic[i] = '0'
    statistic = [_.replace('+','').replace(',','') for _ in statistic]
    # statistic = [statistic[i]='0' for i in range(len(statistic)) if statistic[i] == '']
    return statistic

def all_data(trs):
    """return all statistics containing the latest info on coronavirus worldwide"""
    statistics = [country_data(i) for i in trs]
    for i in range(len(statistics)):
        # statistics[i] = formatdata(statistics[i][:8])
        statistics[i] = formatdata(statistics[i])
    # statistics = [i for i in statistics if str(i[0]).isalpha()]
    # statistics = [_.append('0') for _ in statistics if len(_) !=8]
    return statistics

def database_process(today, all_info):
    # today = '_'.join([str('%02d'%_) for _ in time.localtime()[:3]])
    # create_database = """CREATE DATABASE IF NOT EXISTS covid_19"""
    create_table = """CREATE TABLE IF NOT EXISTS %s(
                      id INT NOT NULL AUTO_INCREMENT,
                      country_region VARCHAR(100) NOT NULL,
                      total_cases INT NOT NULL DEFAULT '0' ,
                      new_cases INT NOT NULL DEFAULT '0',
                      total_death INT NOT NULL DEFAULT '0',
                      new_death INT NOT NULL DEFAULT '0',
                      recovered INT NOT NULL DEFAULT '0',
                      active_cases INT NOT NULL DEFAULT '0',
                      serious_cases INT NOT NULL DEFAULT '0',
                      PRIMARY KEY(id));
                      """ % today
    # cursor.execute(create_database)
    db = pymysql.connect('localhost', 'root', 'tyc1234', 'COVID_19')
    cursor = db.cursor()
    cursor.execute(create_table)
    insert = """INSERT INTO %s(country_region, total_cases,new_cases,total_death,
                new_death, recovered,active_cases, serious_cases) """% today
    for each in all_info:
        values = """VALUES ("%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s");"""% tuple(each)
        sql = insert + values
        # sql = """INSERT INTO %s
        #          (country_region, total_cases,new_cases,total_death,
        #           new_death, recovered,active_cases, serious_cases)
        #          VALUES ("%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s");"""% tuple(each)
        cursor.execute(sql)
    print('成功写入 %d 个国家疫情信息' %len(all_info))
    cursor.connection.commit()
    db.close()

def save_data(name, content):
    print('save data..')
    temp = pickle.dumps(content)
    file_name = name + '_疫情数据.pl'
    # today = '_'.join([str('%02d'%_) for _ in time.localtime()[:3]])
    with open(file_name,'wb') as f:
        f.write(temp)

class DataBase:
    def __init__(self, db, cur):
        self.db = db
        self.cur = cur
        self.tables = self.get_tables(cur)
        self.total_cases_sql = self.select_total_cases(self.tables)
        self.active_cases_sql = self.select_active_cases(self.tables)
        self.data = []
        self.data_combined = []

    @staticmethod
    def get_tables(cur):
        cur.execute('show tables;')
        # temp =
        tables = [_[0] for _ in cur.fetchall()]
        # db.close()
        return tables

    @staticmethod
    def select_total_cases(tables):
        # tables = self.tables
        select_ = 'SELECT %s.country_region'%tables[0]
        for i in tables:
            temp = ', %s.total_cases %s'% (i,i)
            select_ += temp
        from_ = ' FROM %s'%tables[0]
        for i in tables[1:]:
            temp = ', %s' % i
            from_ += temp
        where_ = ' WHERE '
        for i in tables[1:]:
            temp = '%s.country_region=%s.country_region '%(tables[0],i)
            where_ += temp
            if tables.index(i) != len(tables) -1:
                where_ += 'and '
        order_ = 'ORDER BY %s.total_cases DESC;'%tables[-1]
        sql = select_ + from_ + where_ + order_
        # self.total_cases_sql = sql
        return sql

    @staticmethod
    def select_active_cases(tables):
        # tables = self.tables
        select_ = 'SELECT %s.country_region'%tables[0]
        for i in tables:
            temp = ', %s.active_cases %s'% (i,i)
            select_ += temp
        from_ = ' FROM %s'%tables[0]
        for i in tables[1:]:
            temp = ', %s' % i
            from_ += temp
        where_ = ' WHERE '
        for i in tables[1:]:
            temp = '%s.country_region=%s.country_region '%(tables[0],i)
            where_ += temp
            if tables.index(i) != len(tables) -1:
                where_ += 'and '
        order_ = 'ORDER BY %s.active_cases DESC;'%tables[-1]
        sql = select_ + from_ + where_ + order_
        return sql

    def get_data(self,sql):
        self.data.clear()
        self.cur.execute(sql)
        self.data = self.cur.fetchall()
        self.data = [list(_) for _ in self.data]

    def process(self):
        temp = self.tables.copy()
        temp.insert(0, 'country')
        self.data_combined = self.data
        self.data_combined.insert(0, temp)

class SaveExcel:
    def __init__(self):
        self.workbook = 'covid.xlsx'
        self.wb = self.getworkbook()

    @staticmethod
    def getworkbook():
        workbook = 'covid.xlsx'
        if os.path.exists(workbook):
            wb = openpyxl.load_workbook(workbook)
        else:
            wb = openpyxl.Workbook()
        return wb

    def worksheet(self,data, name='data sheet'):
        ws = self.wb.create_sheet(name)
        for i in data:
            ws.append(i)
        self.wb.save(self.workbook)

def main():
    today = '_'.join([str('%02d'%_) for _ in time.localtime()[:3]])
    web = access_web()
    raw_data = get_target(web.text)
    COVID_data = all_data(raw_data)
    save_data(today,COVID_data)
    info_needed = [[_[0],_[1],_[3],_[6]] for _ in COVID_data]
    try:
        database_process(today, info_needed)
    except TypeError as e:
        print('原始数据错误...',e)
    finally:
        return COVID_data

def generate_excel():
    db = pymysql.connect('localhost', 'root', 'tyc1234', 'COVID_19')
    cur = db.cursor()
    excel = SaveExcel()
    covid = DataBase(db, cur)
    covid.get_data(covid.total_cases_sql)
    covid.process()
    excel.worksheet(covid.data_combined, 'total_case')
    covid.get_data(covid.active_cases_sql)
    covid.process()
    excel.worksheet(covid.data_combined, 'active_cases')
    covid.db.close()
    excel.wb.close()

def all_url(trs):
    urls = {}
    for each in trs:
        target = each.find(href=True)
        try:
            country = target.text
            urls[country] = 'https://www.worldometers.info/coronavirus/' + target['href']
        except AttributeError as e:
            error = each.td.text
            print('%s has no links'% error)
    return urls

def extract_data(statistics):
    """pass into the target raw data, which should be a dict
       return a dict containing the extracted info from passed param:
       1、'Number of Infected People':{categories: date, data: numbers}
       2、'coronavirus-death-log':{categories: date, data: numbers}
       3、'coronavirus-cases-log':{categories: date, data: numbers}
       which are all in dict forms as well
       """
    info_extracted = {}
    for k, v in statistics.items():
        pattern1 = re.compile('categories.+?\]')
        categories = pattern1.search(v).group()
        categories = json.loads(categories.split(':')[1])
        # date = [_ + ' 2020' for _ in date]
        categories = [formatdate(_ + ' 2020') for _ in categories]
        pattern2 = re.compile('data.+?\]')
        data = pattern2.search(v).group()
        data = json.loads(data.split(':')[1])
        info_extracted[k] = {'categories': categories, 'data': data}
    return info_extracted

def formatdate(single_date):
    temp = time.strptime(single_date,'%b %d %Y')
    return time.strftime('%Y_%m_%d',temp)

def find_data(page):
    """pass into the text web page of COVID19 of a certain country,
       return a dict containing 3 segments:
       1、'Number of Infected People': strings
       2、'coronavirus-death-log': strings
       3、'coronavirus-cases-log': strings
       which all are raw data,
       in other words, strings
       """
    pattern = re.compile('Highcharts\.chart\([\s\S]*?;')
    data = pattern.findall(page)
    targets = ['Number of Infected People','coronavirus-death-log','coronavirus-cases-log']
    final = {}
    for each in targets:
        for info in data:
            try:
                temp = re.search(each, info).group()
                final[each] = info
            except AttributeError as e:
                pass
    return final

def re_organizeData(data):

    def extract_single(day, statistics):
        date_range = statistics['categories']
        num_range = statistics['data']
        position = date_range.index(day)
        return num_range[position]

    def organize_single_country(one_country):
        """return structured covid info of the given country
           [a, b, c]
           a is total cases
           b is avtive cases
           c is total death
           """
        nation = one_country[0]
        days = one_country[1]['coronavirus-cases-log']['categories']
        covidInfo = {nation:{}}
        for day in days:
            total_cases = extract_single(day, one_country[1]['coronavirus-cases-log'])
            active_cases = extract_single(day,one_country[1]['Number of Infected People'])
            try:
                total_death = extract_single(day,one_country[1]['coronavirus-death-log'])
            except KeyError as e:
                total_death = 0
            covidInfo[nation][day] = [total_cases, active_cases, total_death]
        return covidInfo

    def fetch(day):
        def sort_criteria(element):
            return element[1]

        covid_at_cur_day = {day:[]}
        countries = covidInfo_allCountries.keys()
        for country in countries:
            try:
                data = covidInfo_allCountries[country][day]
            except KeyError as e:
                data = [0, 0, 0]
            finally:
                data.insert(0, country)
                covid_at_cur_day[day].append(data)
        covid_at_cur_day[day].sort(key=sort_criteria, reverse=True)
        return covid_at_cur_day

    covidInfo_allCountries = {}
    for country in  data:
        try:
            covidInfo = organize_single_country(country)
            covidInfo_allCountries.update(covidInfo)
        except KeyError as e:
            e = str(e)
            print(country[0]+ ' >>> ' + e)

    China = [i[1] for i in data if i[0]=='China'][0]
    days = China['coronavirus-cases-log']['categories']
    days_data = {}
    for day in days:
        cur_day_data = fetch(day)
        days_data.update(cur_day_data)
    return days_data

def dump_in(tablename, tabledata):
    # today = '_'.join([str('%02d'%_) for _ in time.localtime()[:3]])
    # create_database = """CREATE DATABASE IF NOT EXISTS covid_19"""
    create_table = """CREATE TABLE IF NOT EXISTS %s(
                      id INT NOT NULL AUTO_INCREMENT,
                      country_region VARCHAR(100) NOT NULL,
                      total_cases INT NOT NULL DEFAULT '0' ,
                      active_cases INT NOT NULL DEFAULT '0',
                      total_death INT NOT NULL DEFAULT '0',
                      PRIMARY KEY(id));
                      """ % tablename
    # cursor.execute(create_database)
    # db = covid.db
    # cursor = covid.cur
    cur.execute(create_table)
    insert = "INSERT INTO %s(country_region, total_cases,active_cases,total_death) "% tablename
    for each in tabledata:
        values = """VALUES ("%s", "%s", "%s", "%s");"""% tuple(each)
        sql = insert + values
        cur.execute(sql)
    cur.connection.commit()
    print('成功写入 %d 个国家疫情信息添加到表单【%s】' %(len(tabledata), tablename))

def updateMysql():
    today = '_'.join([str('%02d'%_) for _ in time.localtime()[:3]])
    url = 'https://www.worldometers.info/coronavirus/'
    web  = access_web(today, url)
    trs = get_target(web)
    urls = all_url(trs)
    all_countries = []
    for country, url in urls.items():
        country_web = access_web(country, url)
        info = find_data(country_web.text)
        single_country = (country, extract_data(info))
        save_data(country, single_country)
        print('【%s】疫情信息保存完毕'%country)
        all_countries.append(single_country)
        time.sleep(2)
    structured_covid_info = re_organizeData(all_countries)
    db = pymysql.connect('localhost', 'root', 'tyc1234', 'COVID_UPDATE')
    cur = db.cursor()
    covid = DataBase(db, cur)
    for day, details in structured_covid_info.items():
        dump_in(day, details)
    return structured_covid_info

def collect_data_into_one(maintable, targetable, exists=False):
    # global exists
    create_table = """CREATE TABLE %s(
                      country_region VARCHAR(100) NOT NULL DEFAULT "unknown",
                      PRIMARY KEY(country_region));
                      """ % maintable
    insert_countries = """INSERT INTO %s(country_region)
                          SELECT country_region FROM %s;
                          """ % (maintable, tables[-1])
    if not exists:
        try:
            cur.execute(create_table)
            print('创建表单 <%s> ' % maintable)
            cur.execute(insert_countries)
            exists = True
        except pymysql.err.InternalError as e:
            exists = True
    if_has_new_countries(maintable,targetable)
    alter = "ALTER TABLE %s ADD %s INT NOT NULL DEFAULT '0'; "% (maintable,targetable)
    try:
        cur.execute(alter)
    except pymysql.err.InternalError as e:
        print('%s 列已存在，尝试更新..'%targetable)
    update_sql = """UPDATE {0} LEFT JOIN {1}
                    ON {0}.country_region={1}.country_region
                    SET {0}.{1}=IFNULL({1}.{0}, 0);
                    """.format(maintable, targetable)
    temp = cur.execute(update_sql)
    cur.connection.commit()
    print('成功将 %d 个国家 %s 的疫情信息到表单【%s】' %(temp, targetable,maintable))
    return exists

def exclude_tables(tables):
    def formatables(tables):
        # cur.execute('show tables;')
        # temp =
        tables = [_[0] for _ in tables]
        # db.close()
        return tables
    tables = formatables(tables)
    pattern = re.compile('\d{2}_\d{2}_\d{2}')
    final = []
    for table in tables:
        try:
            verify = pattern.search(table).group()
            final.append(table)
        except AttributeError as e:
            # print('remove <%s> from tables'% table)
            pass
    return final

def write_in_batch(target, raw_tables):
    exists = False
    for table in raw_tables:
        exists = collect_data_into_one(target, table, exists)

def organize_all_tables_into_one():
    """connect to database and generate sql to collect
       statistics of <total_cases>,<active_cases> and <total_death>
       from each table of respective date,
       then produce 3 seperate tables to store the collected statistics
       ---TABLE <total_cases>: all total cases from different date
       ---TABLE <active_cases>: all active cases from different date
       ---TABLE <total_death>: all total death data from different date"""
    db = pymysql.connect('localhost', 'root', 'tyc1234', 'COVID_UPDATE')
    cur = db.cursor()
    cur.execute('show tables;')
    tables = cur.fetchall()
    tables = exclude_tables(tables)
    write_in_batch('total_cases', tables)
    write_in_batch('active_cases', tables)
    write_in_batch('total_death', tables)
    db.close()
    return tables

def if_has_new_countries(target_table,latest_table):
    sql = """INSERT INTO {target_table}(country_region)
             SELECT latest
                FROM
                    (SELECT
                        t1.country_region target,
                        t2.country_region latest
                    FROM
                        {target_table} t1
                        RIGHT JOIN {latest_table} t2
                    ON
                        t1.country_region=t2.country_region
                    WHERE
                        t1.country_region IS NULL)
             AS ref;""".format(target_table=target_table, latest_table=latest_table)
    res = cur.execute(sql)
    if res:
        print('%d newly added countries for table <%s>.' % (res, target_table))

def change_name(tablename, column1='total_death', column2='total_death'):
    """定义这个函数纯粹只是为了修改之前的一些错误，death后面加了一个s
       但是death是不可数啊，仙人板板"""
    sql = """ALTER TABLE
                {tablename}
            CHANGE
                {column1}
                {column2} INT NOT NULL DEFAULT '0';
          """.format(tablename=tablename,column1=column1,column2=column2)
    cur.execute(sql)


if __name__ == "__main__":
    db = pymysql.connect('localhost', 'root', 'tyc1234', 'COVID_UPDATE')
    cur = db.cursor()
    # cur.execute('show tables;')
    # tables = exclude_tables(cur.fetchall())
    url = 'https://www.worldometers.info/coronavirus/'
    today = '_'.join([str('%02d'%_) for _ in time.localtime()[:3]])
    web = access_web(today, url)
    raw_data = get_target(web.text)
    COVID_data = all_data(raw_data)
    save_data(today,COVID_data)
    info_needed = [[_[0],_[1],_[6],_[3]] for _ in COVID_data]
    exists = True
    targetables = ['total_cases','active_cases','total_death']
    try:
        dump_in(today, info_needed)
        cur.execute('show tables;')
        tables = exclude_tables(cur.fetchall())
        # execute1 = [if_has_new_countries(_,tables[-1]) for _ in targetables]
        execute = [collect_data_into_one(_, today, exists) for _ in targetables]
    except TypeError as e:
        print('原始数据错误...',e)
